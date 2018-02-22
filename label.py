# -*- coding: utf-8 -*-
"""
Created on Mon Feb 12 14:09:33 2018

@author: gwu


cell labels
Table should look roughly like:
    ______________________________________________________________
    |Image | Cell # | ROI coord | labels/characteristics | ... TBA|
    |------------------------------------------------------------ |
    |      |        |           |                        |        |
    |      |        |           |                        |        | 

"""

from openpyxl import Workbook

class cellChars(cellNum, descr):
    
    def __init__():
        wb = Workbook()
        ws = wb.active
        wb.save("sample.xlsx")

        ws['A1'] = 'Cell #'
        ws['B1'] = 'ROI Coord'
        ws['C1'] = 'Image'
        ws['D1'] = 'Undifferentiated'
        ws['E1'] = 'Differentiated'
        ws['F1'] = 'Damaged'
        
        curr = 'A2'
        
    def addCellNum():
        for row in ws.iter_rows(min_row = 2, max_col = 1, max_row = 10000):
            if cell in row == None:
                ws[cell] = cellNum
            return
        for col in ws.iter_rows(min_row = 2, max_col = 10000, max_row = 2):
            if ws[vol] == None:
                ws['A'+ row] = descr
                curr = 'A' +row
            return
        for col in range(,):
            if ws[col + ]
    

    # Create an new Excel file and add a worksheet.
    #EXCEL FILE CURRENTLY SAVES TO PATH:
    #C:\Users\gwu\git\SBP-project
    workbook = xlsxwriter.Workbook('test.xlsx')
    worksheet = workbook.add_worksheet()
    
    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})
    
    # Headers
    locked = workbook.add_format()
    locked.set_locked(True)
    worksheet.write('A1', 'Image', bold)
    worksheet.write('B1', 'Cell #', bold)
    
    #Cell center coord, length, width
    worksheet.write('C1', 'Cell Coord', bold)
    worksheet.write('D1', 'Labels', bold)
    
    # Insert an image.
    #TODO image from different folder?
    worksheet.insert_image('A1', 'logo.png')