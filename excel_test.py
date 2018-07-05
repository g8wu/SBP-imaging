# -*- coding: utf-8 -*-
"""
Created on Fri Feb 09 14:59:48 2018

@author: gwu

Append to col 52 of Podocytes_2_12.xlsx
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from Tkinter import Tk
from tkFileDialog import askopenfilename


if __name__ == '__main__':
    
    Tk().withdraw()     #keeps blank tk window from popping up
    filename = askopenfilename()
    wb = load_workbook(filename = filename)
    sheet = wb.active
    
    boundBox = sheet['P']
    
    
    
    #TODO add tags column
    
    #TODO read ROI location
        # Data can be assigned directly to cells
        #ws['A1'] = 42
        
        # Rows can also be appended
        #ws.append([1, 2, 3])
        
        # Save the file
        #wb.save("sample.xlsx")

